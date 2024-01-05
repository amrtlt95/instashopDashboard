from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
from itertools import zip_longest
import sys
import csv

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
import time
import time
import psutil
from datetime import datetime
import os
import random

def loginDashBoard(driver):
    driver.get("https://instashop.ae/dashboard/index.html#/admin")  # open the dashboard
    try:  # handle the login screen if it appeared
        # try to find the element by its CSS selector
        element = WebDriverWait(driver, 5).until(
            EC.invisibility_of_element((By.ID, "login"))
        )
        productsButton = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div[8]/div[1]/div/div[2]/ul/li[2]/a")))
        productsButton.click()

    except:
        print("Logged out\n")
        userName = driver.find_element(By.CSS_SELECTOR, "#login > form > div:nth-child(1) > div > input")
        password = driver.find_element(By.CSS_SELECTOR, "#login > form > div:nth-child(2) > div > input")
        loginButton = driver.find_element(By.CSS_SELECTOR, "#login > form > div.submit.text-right > button")
        userName.send_keys("amr.khalil@instashop.ae")
        password.send_keys("12345@Aab")
        loginButton.click()
        element = WebDriverWait(driver, 5).until(
            EC.invisibility_of_element((By.ID, "login"))
        )
        productsButton = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div[8]/div[1]/div/div[2]/ul/li[2]/a")))
        productsButton.click()


def slowConnection():
    print("The site is slow , kindly check network connection and try again\n")

def is_file_opened(file_path):
    try:
        file_descriptor = os.open(file_path, os.O_RDWR | os.O_CREAT)
        os.close(file_descriptor)
        return False
    except OSError:
        return True

def getClientName(driver):
    clientName = input("Enter client name\n")
    return clientName #

def elementWait(driver,timeOut,byWhat,selector):
    try:
       if byWhat=='id':
           element = WebDriverWait(driver, timeOut).until(
               EC.visibility_of_element_located((By.ID,selector))

        )
       elif byWhat=='css selector':
           element = WebDriverWait(driver, timeOut).until(
               EC.visibility_of_element_located((By.CSS_SELECTOR, selector))
           )
       elif byWhat=='class name':
           element = WebDriverWait(driver, timeOut).until(
               EC.visibility_of_element_located((By.CLASS_NAME, selector))
           )

       elif byWhat=='x path':
           element = WebDriverWait(driver, timeOut).until(
               EC.visibility_of_element_located((By.XPATH, selector))
           )

    except:
        slowConnection()
        driver.quit()

def elementsWait(driver,timeOut,byWhat,selector):
    try:
       if byWhat=='id':
           element = WebDriverWait(driver, timeOut).until(
               EC.visibility_of_all_elements_located((By.ID,selector))
        )
       elif byWhat=='css selector':
           element = WebDriverWait(driver, timeOut).until(
               EC.visibility_of_all_elements_located((By.CSS_SELECTOR, selector))
           )
       elif byWhat == 'class name':
           element = WebDriverWait(driver, timeOut).until(
               EC.visibility_of_all_elements_located((By.CLASS_NAME, selector))
           )
    except:
        slowConnection()
        driver.quit()

def clientSelect(driver,clientName):
    driver.find_element(By.CSS_SELECTOR,"#clientSelect_chosen > a").click()
    clientNameTextBox=driver.find_element(By.CSS_SELECTOR,"#clientSelect_chosen > div > div > input")
    clientNameTextBox.send_keys(clientName)
    try:
        WebDriverWait(driver,timeout=3).until(
        EC.visibility_of_element_located((By.CLASS_NAME,"no-results"))
        )
        print("Wrong Client Name , please try again")
        clientNameTextBox.clear()
    except:
        clientNameTextBox.send_keys(Keys.RETURN)
def getBusniessType(driver):
    businessType = int(input(
        "What is the business type\n1) Pharmacy\n2) Pet Shop\n3) Specialty\n4) Grocery\n"))  # asks for business type so that it selects the report's template
    driver.maximize_window()
    return businessType

def eodFilter(driver):

    #driver.find_element(By.CSS_SELECTOR,
     #                   "#page-wrapper > div:nth-child(4) > div:nth-child(2) > div:nth-child(1)").click()  # Enable Filterz
    enabledFilter=WebDriverWait(driver,50).until(
        EC.element_to_be_clickable((By.XPATH,"/html/body/div[1]/div[9]/div[146]/div[4]/div[1]/div[1]"))
    )
    enabledFilter.click()
    time.sleep(3)
    driver.find_element(By.CSS_SELECTOR,
                        "#page-wrapper > div:nth-child(6) > div.col-xs-12 > div > div:nth-child(2) > button").click()  # load products

def oosFilters(driver):
    waitForLoading(driver)
    enabledFilter = WebDriverWait(driver, 50).until(
        EC.visibility_of_element_located(
            (By.XPATH, "/html/body/div[1]/div[9]/div[145]/div[4]/div[1]/div[1]"))
    )
    oosFilter= WebDriverWait(driver, 50).until(
        EC.visibility_of_element_located(
            (By.XPATH, "/html/body/div[1]/div[9]/div[145]/div[4]/div[3]/div[2]"))
    )
    noTimingsFilter = WebDriverWait(driver, 50).until(
        EC.visibility_of_element_located(
            (By.XPATH,  "/html/body/div[1]/div[9]/div[145]/div[4]/div[4]/div[11]"))
    )

    hasBarcodeFilter = WebDriverWait(driver, 50).until(
        EC.visibility_of_element_located(
            (By.XPATH, "/html/body/div[1]/div[9]/div[145]/div[4]/div[5]/div[4]"))
    )

    noWapFilter = WebDriverWait(driver, 50).until(
        EC.visibility_of_element_located(
            (By.XPATH,  "/html/body/div[1]/div[9]/div[145]/div[4]/div[2]/div[9]"))
    )
    hasBrandFilter=WebDriverWait(driver, 50).until(
        EC.visibility_of_element_located(
            (By.XPATH,  "/html/body/div[1]/div[9]/div[145]/div[4]/div[4]/div[4]"))
    )
    enabledFilter.click()
    oosFilter.click()
    noTimingsFilter.click()
    hasBrandFilter.click()
    hasBarcodeFilter.click()
    noWapFilter.click()

def hasWapClick(driver):
    #driver.find_element(By.CSS_SELECTOR,
     #                   "#page-wrapper > div:nth-child(4) > div:nth-child(2) > div:nth-child(1)").click()  # Enable Filterz
    hasWapFilter=WebDriverWait(driver,50).until(
        EC.element_to_be_clickable((By.XPATH,"/html/body/div[1]/div[9]/div[146]/div[4]/div[2]/div[8]"))
    )
    hasWapFilter.click()
    time.sleep(3)
    driver.find_element(By.CSS_SELECTOR,
                        "#page-wrapper > div:nth-child(6) > div.col-xs-12 > div > div:nth-child(2) > button").click()  # load products


def setGoogleApi(jsonPath):
    # Set up the credentials for accessing the Google Sheet API
    scope = ["https://spreadsheets.google.com/feeds"]
    credentials = ServiceAccountCredentials.from_json_keyfile_name(jsonPath, scope)
    return gspread.authorize(credentials)

def openGoogleSheetsWorkbookByKey(gc,workbookKey):
    source_spreadsheet = gc.open_by_key(workbookKey)  # google template
    return source_spreadsheet




def selectGoogleTemplateSheet(driver,businessType,sourceGoogleWorkBook):
    if businessType == 1:  # then the business is a pharmacy , we will open the pharmacy sheet that will be used to copy all the formatting
        sourceGoogleSheet = sourceGoogleWorkBook.worksheet("Pharmacy")
        return sourceGoogleSheet

    elif businessType == 2:  # then the business is a petshop , we will open the petshop sheet that will be used to copy all the formatting
        # Open the Pet Shop worksheet
        sourceGoogleSheet = sourceGoogleWorkBook.worksheet("Petshop")
        return sourceGoogleSheet
    elif businessType ==3:
        #open the Specialty worksheet
        sourceGoogleSheet = sourceGoogleWorkBook.worksheet("Specialty")#Specialty
        return sourceGoogleSheet
    elif businessType ==4:
        #open the Specialty worksheet
        sourceGoogleSheet = sourceGoogleWorkBook.worksheet("Grocery")
        return sourceGoogleSheet

    else:
        print("Invalid business type")
        driver.quit()


def selectExcelFormattingSheet(driver,businessType,sourceExcelWorkBook):
    if businessType == 1:  # then the business is a pharmacy , we will open the pharmacy sheet that will be used to copy all the formatting
        sourceExcelsheet = sourceExcelWorkBook["Pharmacy"]  # worksheet that we will copy format from
        return sourceExcelsheet

    elif businessType == 2:  # then the business is a petshop , we will open the petshop sheet that will be used to copy all the formatting
        # Open the Pet Shop worksheet
        sourceExcelsheet = sourceExcelWorkBook["Petshop"]  # worksheet that we will copy format from
        return sourceExcelsheet
    elif businessType ==3:
        sourceExcelsheet = sourceExcelWorkBook["Specialty"]
        return sourceExcelsheet

    elif businessType == 4:
        sourceExcelsheet = sourceExcelWorkBook["Grocery"]


        return sourceExcelsheet
    else:
        print("Invalid business type")
        driver.quit()

def copyAllGoogleSheetsToExcel(googleSheet, excelSheet):
    for row in googleSheet.get_all_values():  # This will copy all the text from google sheets
        excelSheet.append(row)

def copyCellFormating(sourceExcelSheet,destExcelSheet):
    # Copy formatting, values, and limited to column 18 and row 100
    for row_index, row in enumerate(sourceExcelSheet.iter_rows(), start=1):
        if row_index > 100:
            break  # Stop copying after reaching row 100

        for column_index, cell in enumerate(row, start=1):
            if column_index > 18:
                break  # Stop copying after reaching column 18

            dest_cell = destExcelSheet.cell(row=row_index, column=column_index)

            # Copy the cell value
            dest_cell.value = cell.value

            # Copy the cell formatting
            dest_cell.number_format = cell.number_format
            dest_cell.font = copy(cell.font)
            dest_cell.alignment = copy(cell.alignment)
            dest_cell.border = copy(cell.border)
            dest_cell.fill = copy(cell.fill)
            dest_cell.protection = copy(cell.protection)

    # Copy column widths up to column 18
    for column_index in range(1, 19):
        column_letter = get_column_letter(column_index)
        destExcelSheet.column_dimensions[column_letter].width = sourceExcelSheet.column_dimensions[column_letter].width

    # Copy row heights up to row 100
    for row_index in range(1, 101):
        destExcelSheet.row_dimensions[row_index].height = sourceExcelSheet.row_dimensions[row_index].height

    # Copy merged cells
    for merged_range in sourceExcelSheet.merged_cells.ranges:
        destExcelSheet.merge_cells(merged_range.coord)


def iter_rows(ws): # this to get rid of none values in the category list of lists
    for row in ws.iter_rows():

        yield [cell.value for cell in row]

def getExcelCategories(busnissType,excelSheet):
    listLists = [[]]
    tempList = []
    categoriesList = []

    listLists = iter_rows(excelSheet)
    for list in listLists:
        for element in list:
            if element != None:
                element = element.rstrip()
                element = element.lstrip()
                tempList.append(element)
    if busnissType==1:
        categoriesList = tempList[8:]
    elif busnissType == 2:
        categoriesList = tempList[8:]
    elif busnissType==3:
        categoriesList = tempList[7:]


    return categoriesList


def replaceCellValue(worksheet,oldValue,newValue):
    # Iterate over rows
    # Iterate over rows, starting from the second row (min_row=2)
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        for column_index, cell_value in enumerate(row, start=1):
            if cell_value == oldValue:
                # Modify the value in the cell
                worksheet.cell(row=row_index, column=column_index, value=newValue)

                # Break out of the inner loop
                break

        # If the value was replaced, break out of the outer loop as well
        else:
            continue  # Continue to the next row
        break  # Break out of the outer loop

def waitForLoading(driver):
        WebDriverWait(driver, 60).until(
            EC.invisibility_of_element((By.CLASS_NAME, "fa-fw"))
        )


def waitForFakeLoading(driver):
    try:
        WebDriverWait(driver, 120).until(
            EC.invisibility_of_element((By.XPATH, " /html/body/div/div[10]"))
        )
    except:
        slowConnection()
        driver.quit()
        exit()

def waitForProgressBarInInfo(driver):
    try:
        WebDriverWait(driver,120).until(

            EC.invisibility_of_element((By.XPATH,"/html/body/div[1]/div[9]/div[13]/div/div/div/div/div[2]/div/div[1]/md-progress-linear"))
        )
    except:
        slowConnection()
        exit()
def removeSpacesInSheet(excelSheet):
    merged_ranges = excelSheet.merged_cells.ranges

    # Iterate through rows in the specified range
    for row_index, row in enumerate(excelSheet.iter_rows(min_row=5, max_row=100, min_col=2, max_col=16), start=5):
        for column_index, cell in enumerate(row, start=2):
            is_merged = False

            # Check if the cell is part of a merged range
            for merged_range in merged_ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    break

            if not is_merged:
                if cell.value is not None and isinstance(cell.value, str):
                    stripped_value = cell.value.strip()
                    excelSheet.cell(row=row_index, column=column_index, value=stripped_value)


def fillMergedCellWithInfo(excelSheet, start_row, end_row, start_col, end_col, value):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = excelSheet.cell(row=row, column=col)

            # Check if the cell's coordinate is within a merged range
            is_merged = False
            for merged_range in excelSheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    break

            if is_merged:
                # Unmerge the cell
                start_cell = excelSheet.cell(row=start_row, column=start_col)
                end_cell = excelSheet.cell(row=end_row, column=end_col)
                cell_range = f"{start_cell.coordinate}:{end_cell.coordinate}"
                excelSheet.unmerge_cells(cell_range)

                # Update the value of the first cell in the merged range
                first_cell = excelSheet[start_cell.coordinate]
                first_cell.value += value

                # Merge the cell again
                excelSheet.merge_cells(cell_range)
                break

            else:
                # Update the value of the non-merged cell
                cell.value += value
        break




def getTeamLeader():
    teamLeader=input("Enter the team leader name\n")
    return teamLeader
def getTeamMembers():
    # Get input from the user
    input_string = input("Enter Team members' names seperated by dash '-' :\nAmr-Hossam-Ahmed Abd Al Azim\n")

    # Split the input string into a list of strings
    input_list = input_string.split(sep="-")
    return input_list
def mergeTeamMembers(teamMembersList):
    teamMembers=teamMembersList
    mergedTeamMembers=""
    for member in teamMembers:
        mergedTeamMembers+= (member+",\n")
    return mergedTeamMembers[:-2]
def nonGroceryInfo(excelSheet,startRow,endRow,startCol,endCol,value):

    fillMergedCellWithInfo(excelSheet,startRow,endRow,startCol,endCol,value)


def generateMemberInCell(max,list,excelSheet,row,col):
    randomNumber = random.randint(0, max)
    excelSheet.cell(row, col).value = list[randomNumber]



def askingForToday(busniessType):
    if busniessType == 4:
        howManyDays = int(input("How many days did you take in the shop\n"))
        if howManyDays > 1:
            firstDate = input("Provide the 'day 1' date. For example : 18/12/1996\n")
            lastDate = input("Provide the last date. For example : 18/12/1996\n")
            datesArr=[firstDate,lastDate]
            return datesArr

    answer = input("Is the report today?\n(y/n): ")

    if answer=="n":
        answer = input("Provide the date. For example : 18/12/1996\n")
        return answer
    else:
        return answer

def findCorrectVertical(driver , bussniessType):
    fullDashboardCategories = None
    if bussniessType == 1:
        fullDashboardCategories = WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located(
                (By.XPATH, "/html/body/div[1]/div[9]/div[146]/div[7]/div[1]/ul/div[4]"))
        )
    elif bussniessType == 2:
        fullDashboardCategories = WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located(
                (By.XPATH, "/html/body/div[1]/div[9]/div[145]/div[7]/div[1]/ul/div[5]"))
        )
    elif bussniessType == 3:

        fullDashboardCategories = WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located(
                (By.XPATH, "/html/body/div[1]/div[9]/div[146]/div[7]/div[1]/ul/div[12]/div"))
        )

    elif bussniessType == 4:

        fullDashboardCategories = WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located(
                (By.XPATH, "/html/body/div[1]/div[9]/div[146]/div[7]/div[1]/ul/div[1]"))
        )
    else:
        exit(0)
    return fullDashboardCategories



# All elements in the dashboard
def productsSearchButton(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[146]/div[6]/div[1]/div/div[2]/button")

def allCategoriesButton(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[146]/div[7]/div[1]/button[3]")


def duplicatedElements():
    return[]

def loopUntilProductsFound(driver):
    try:
        waitForLoading(driver)
    except:
        productsSearchButton(driver).click()
        loopUntilProductsFound(driver)


def returnProductsSearchTextBox(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[146]/div[6]/div[1]/div/div[1]/div[1]/div/input")


def returnElementInfoButton(driver):
    return driver.find_element(By.CSS_SELECTOR,"#allProducts > div:nth-child(3) > div > div.product-title > div.statusProducts > div.cursor-pointer.solidonhover.d-inline-block")
def returnPluFromInfoBulletButton(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[13]/div/div/div/div/div[2]/div/md-radio-group/md-radio-button[12]")
def returnNextButtonInInfo(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[13]/div/div/div/div/div[3]/button[2]")

def returnApprovedPLUOrBarCodes(driver):
    return driver.find_elements(By.CSS_SELECTOR,"tr:nth-child(2) > .centered:nth-child(2) > span > div")
def returnCloseButtonInPluSectionInInfoMenu(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[13]/div/div/div/div/div[3]/button[2]")

def returnBackButtonInInfoMenu(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[13]/div/div/div/div/div[3]/button[1]")

def returnBarcodesFromInfoBulletButton(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[13]/div/div/div/div/div[2]/div/md-radio-group/md-radio-button[2]")
def checkIfItemHasApprovedPlu(driver):
    pluBox= driver.find_element(By.XPATH,"//span[text()='PLUs']")
    approvedPluAttribute= pluBox.get_attribute("ng-if")
    return "item.plu.length > 0" in approvedPluAttribute

def checkIfItemIsEnabled(driver):
    enabledStatusBox=driver.find_element(By.XPATH,"//div[@ng-click='searchForConflictsBeforeToggling(item)']")
    enabledStatusBoxAttribute=enabledStatusBox.get_attribute("class")
    return "clientOption active" in enabledStatusBoxAttribute


def returnBarcodeAndPLUBoxButton(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[146]/div[7]/div[2]/div[3]/div/div[3]/div[7]/div[3]/img")
def returnBarcodesAndPLUsAfterClickOnTheBox(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[113]/div/div/div/div/div[2]")


def returnEnabledProducts(driver):
    return driver.find_elements(By.CSS_SELECTOR,'div#allProducts > div div.admin-product')

def returnNextBottomDashboardButton(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[146]/div[8]/div/button[2]")

def returnProductSearchBox(driver):
    return driver.find_element(By.XPATH,"/html/body/div[1]/div[9]/div[146]/div[6]/div[1]/div/div[1]/div[1]/div/input")
